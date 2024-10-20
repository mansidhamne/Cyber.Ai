import torch
from transformers import AutoModelForCausalLM, AutoTokenizer, pipeline
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
import json
from datetime import datetime
from sentence_transformers import SentenceTransformer
import numpy as np
from enum import Enum
from typing import List, Dict, Optional, Tuple
import json
from datetime import datetime
from sentence_transformers import SentenceTransformer
import numpy as np
from enum import Enum
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load model and tokenizer
model = AutoModelForCausalLM.from_pretrained(
    "microsoft/Phi-3-mini-4k-instruct",
    device_map="cuda",
    torch_dtype="auto",
    trust_remote_code=True,
)
tokenizer = AutoTokenizer.from_pretrained("microsoft/Phi-3-mini-4k-instruct")

# Create a pipeline
pipe = pipeline(
    "text-generation",
    model=model,
    tokenizer=tokenizer,
    return_full_text=False,
    max_new_tokens=500,
    do_sample=False,
)

class RiskLevel(Enum):
    LOW = "Low"
    MEDIUM = "Medium"
    HIGH = "High"
    CRITICAL = "Critical"

    def to_dict(self):
        return {'level': self.name, 'description': self.value}

@dataclass
class SecurityDomain:
    name: str
    keywords: List[str]
    best_practices: List[str]
    risk_indicators: List[str]
    follow_up_templates: List[str]

@dataclass
class SecurityResponse:
    question: str
    answer: str
    domain: str
    risk_score: float
    risk_level: RiskLevel
    timestamp: str
    recommendations: List[str]
    reasoning: str

class DynamicSecurityConversationManager:
    def __init__(self, config_path: str = "/content/knowledge_base.json"):
        self.conversation_history: List[SecurityResponse] = []
        self.risk_scores: Dict[str, List[float]] = {}
        self.encoder = SentenceTransformer('all-mpnet-base-v2')
        self.question_count = {domain: 0 for domain in self._load_domains(config_path)}

        # Load configuration from JSON
        with open(config_path, 'r') as f:
            config = json.load(f)

        # Initialize security domains from configuration
        self.security_domains = {}
        for domain_name, domain_data in config['domains'].items():
            self.security_domains[domain_name] = SecurityDomain(
                name=domain_data['name'],
                keywords=domain_data['keywords'],
                best_practices=domain_data['best_practices'],
                risk_indicators=domain_data['risk_indicators'],
                follow_up_templates=domain_data['follow_up_templates']
            )

        # Store risk thresholds
        self.domain_risk_thresholds = config['risk_thresholds']

    def _load_domains(self, config_path: str) -> List[str]:
        """
        Load security domains from JSON configuration.
        """
        with open(config_path, 'r') as f:
            config = json.load(f)
        return list(config['domains'].keys())

    def process_response(self, question: str, answer: str) -> Dict:
        """
        Process a security-related response and generate analysis.

        Args:
            question: The security question asked
            answer: The response provided

        Returns:
            Dict containing analysis results
        """
        # Identify the security domain
        domain, domain_confidence = self._identify_domain(question, answer)

        # Score the response
        risk_score, risk_level, reasoning = self._analyze_response(domain, answer)

        # Generate recommendations
        recommendations = self._generate_recommendations(domain, risk_score, answer)

        # Create response object
        response = SecurityResponse(
            question=question,
            answer=answer,
            domain=domain,
            risk_score=risk_score,
            risk_level=risk_level,
            timestamp=datetime.now().isoformat(),
            recommendations=recommendations,
            reasoning=reasoning
        )

        # Update history and scores
        self.conversation_history.append(response)
        if domain not in self.risk_scores:
            self.risk_scores[domain] = []
        self.risk_scores[domain].append(risk_score)

        # Generate next question
        next_question = self._generate_follow_up(domain, answer, risk_score)

        self.question_count[domain] += 1

        # Change domain if the threshold is reached
        if self.question_count[domain] >= 2:  # Change this threshold as needed
            domain = self._change_domain(domain)  # Logic to switch to another domain
            self.question_count[domain] = 0  # Reset counter after switching

        return {
            "current_response": response.__dict__,
            "next_question": next_question,
            "domain_confidence": domain_confidence,
            "risk_assessment": self._calculate_risk_assessment()
        }

    def _identify_domain(self, question: str, answer: str) -> Tuple[str, float]:
        """
        Identify the most relevant security domain for the response.
        """
        combined_text = f"{question} {answer}"
        text_embedding = self.encoder.encode([combined_text])[0]

        best_match = None
        best_score = -1

        for domain_name, domain in self.security_domains.items():
            domain_context = " ".join(domain.keywords + domain.best_practices)
            domain_embedding = self.encoder.encode([domain_context])[0]
            similarity = np.dot(text_embedding, domain_embedding)

            if similarity > best_score:
                best_score = similarity
                best_match = domain_name

        return best_match, float(best_score)

    def _analyze_response(self, domain: str, answer: str) -> Tuple[float, RiskLevel, str]:
        """
        Analyze the response for risk scoring.
        """
        domain_info = self.security_domains[domain]
        answer_embedding = self.encoder.encode([answer])[0]

        # Calculate similarity with best practices
        best_practices_embedding = self.encoder.encode(domain_info.best_practices)
        best_practices_similarities = np.dot(answer_embedding, best_practices_embedding.T)
        best_practices_score = float(np.mean(best_practices_similarities))

        # Calculate similarity with risk indicators
        risk_indicators_embedding = self.encoder.encode(domain_info.risk_indicators)
        risk_indicator_similarities = np.dot(answer_embedding, risk_indicators_embedding.T)

        risk_indicator_weights = [1.5 if "no" in ind.lower() or "lack" in ind.lower() else 1.0
                                for ind in domain_info.risk_indicators]
        weighted_risk_score = float(np.dot(risk_indicator_similarities, risk_indicator_weights) / sum(risk_indicator_weights))

        # Calculate overall risk score
        risk_score = (best_practices_score + (1 - weighted_risk_score)) / 2

        # Get domain-specific threshold
        threshold = self.domain_risk_thresholds.get(domain, 0.7)

        # Determine risk level
        if risk_score >= threshold:
            risk_level = RiskLevel.LOW
        elif risk_score >= threshold - 0.2:
            risk_level = RiskLevel.MEDIUM
        elif risk_score >= threshold - 0.4:
            risk_level = RiskLevel.HIGH
        else:
            risk_level = RiskLevel.CRITICAL

        # Generate reasoning
        reasoning = self._generate_reasoning(domain, best_practices_score, weighted_risk_score)

        return risk_score, risk_level, reasoning

    def _generate_recommendations(self, domain: str, risk_score: float, answer: str) -> List[str]:
        """
        Generate recommendations based on the response analysis.
        """
        domain_info = self.security_domains[domain]
        answer_embedding = self.encoder.encode([answer])[0]

        recommendations = []
        practices_embeddings = self.encoder.encode(domain_info.best_practices)
        similarities = np.dot(answer_embedding, practices_embeddings.T)

        for i, similarity in enumerate(similarities):
            if similarity < 0.5:
                recommendations.append(f"Consider implementing: {domain_info.best_practices[i]}")

        return recommendations

    def _generate_follow_up(self, domain: str, answer: str, risk_score: float) -> str:
        """
        Generate a follow-up question based on the context.
        """
        domain_info = self.security_domains[domain]

        if risk_score < 0.5:
            templates = [t for t in domain_info.follow_up_templates
                       if any(word in t.lower() for word in ["how", "what measures", "plan"])]
        else:
            templates = [t for t in domain_info.follow_up_templates
                       if any(word in t.lower() for word in ["monitor", "review", "audit"])]

        if not templates:
            templates = domain_info.follow_up_templates

        template = np.random.choice(templates)
        return template.format(topic=domain_info.name.replace("_", " "))

    def _generate_reasoning(self, domain: str, best_practices_score: float, risk_indicator_score: float) -> str:
        """
        Generate explanation for the risk assessment.
        """
        return (f"Domain: {domain}\n"
                f"Alignment with best practices: {best_practices_score:.2f}\n"
                f"Presence of risk indicators: {risk_indicator_score:.2f}")

    def _calculate_risk_assessment(self) -> Dict:
        """
        Calculate overall risk assessment across all domains.
        """
        assessment = {
            "overall_risk_score": 0.0,
            "domain_scores": {},
            "timestamp": datetime.now().isoformat()
        }

        if not self.risk_scores:
            return assessment

        for domain, scores in self.risk_scores.items():
            domain_score = sum(scores) / len(scores)
            assessment["domain_scores"][domain] = {
                "score": domain_score,
                "risk_level": (RiskLevel.LOW if domain_score >= 0.8
                      else RiskLevel.MEDIUM if domain_score >= 0.6
                      else RiskLevel.HIGH if domain_score >= 0.4
                      else RiskLevel.CRITICAL).value
            }
            assessment["overall_risk_score"] += domain_score

        assessment["overall_risk_score"] /= len(self.risk_scores)
        return assessment

    def generate_report(self) -> Dict:
        """
        Generate comprehensive security assessment report.
        """
        return {
            "assessment_summary": self._calculate_risk_assessment(),
            "conversation_history": [self._serialize_response(response) for response in self.conversation_history],
            "recommendations": self._compile_overall_recommendations(),
            "timestamp": datetime.now().isoformat()
        }

    def _serialize_response(self, response: SecurityResponse) -> Dict:
        """
        Serialize SecurityResponse for JSON output.
        """
        return {
            "question": response.question,
            "answer": response.answer,
            "domain": response.domain,
            "risk_score": response.risk_score,
            "risk_level": response.risk_level.value,  # Convert enum to string
            "timestamp": response.timestamp,
            "recommendations": response.recommendations,
            "reasoning": response.reasoning
        }

    def _compile_overall_recommendations(self) -> Dict[str, List[str]]:
        """
        Compile recommendations by domain.
        """
        domain_recommendations = {}
        for response in self.conversation_history:
            if response.domain not in domain_recommendations:
                domain_recommendations[response.domain] = []
            domain_recommendations[response.domain].extend(response.recommendations)

        # Remove duplicates while preserving order
        for domain in domain_recommendations:
            domain_recommendations[domain] = list(dict.fromkeys(domain_recommendations[domain]))

        return domain_recommendations


    # def _change_domain(self, current_domain: str) -> str:
    #     """
    #     Change the current domain based on recent conversations.
    #     """
    #     current_domain = self.conversation_history[-1].domain
    #     recent_domains = {response.domain for response in self.conversation_history[-self.question_count[current_domain]:]}

    #     # Get all available domains
    #     all_domains = list(self.security_domains.keys())

    #     # Filter out the recent domains from the available domains
    #     available_domains = [domain for domain in all_domains if domain not in recent_domains and domain != current_domain]

    #     # If there are no available domains left, reset the recent domains tracking
    #     if not available_domains:
    #         available_domains = all_domains  # Reset to all domains if all have been discussed

    #     # Select a new domain randomly from the available ones
    #     new_domain = np.random.choice(available_domains)

    #     return new_domain
    def _change_domain(self, current_domain: str) -> str:
        recent_domains = {response.domain for response in self.conversation_history[-self.question_count[current_domain]:]}
        all_domains = list(self.security_domains.keys())
        available_domains = [domain for domain in all_domains if domain not in recent_domains and domain != current_domain]

        if not available_domains:
            available_domains = all_domains  # Reset to all domains if all have been discussed

        new_domain = np.random.choice(available_domains)
        self.question_count[new_domain] = 0  # Reset the question count for the new domain

        return new_domain


def get_valid_input(prompt: str, input_type: str = "question") -> str:
    """
    Helper function to get valid input from user.
    Keeps prompting until a non-empty response is provided.

    Args:
        prompt: The prompt to show to the user
        input_type: Type of input ("question" or "answer") for customized messages

    Returns:
        str: Valid user input
    """
    while True:
        user_input = input(prompt).strip()

        # Handle exit command
        if user_input.lower() == 'exit':
            return 'exit'

        # Check for empty input
        if not user_input:
            if input_type == "question":
                print("Question cannot be empty. Please enter a valid question.")
            else:
                print("Response cannot be empty. Please provide a detailed answer.")
            continue

        # Check for minimum length (optional)
        if len(user_input.split()) < 2 and input_type == "answer":
            print("Please provide a more detailed response (at least a few words).")
            continue

        return user_input

def run_security_assessment():
    """
    Run an interactive security assessment conversation with improved error handling
    and input validation.
    """
    try:
        # Initialize the conversation manager
        manager = DynamicSecurityConversationManager()

        # Print welcome message and instructions
        print("\n=== Security Assessment Conversation ===")
        print("---------------------------------------")
        print("Instructions:")
        print("- Provide detailed responses for better analysis")
        print("- Type 'exit' at any time to end the assessment and generate report")
        print("- Empty responses are not accepted")
        print("---------------------------------------\n")

        question_count = 1
        next_question = "What is the compliance protocol you follow?"

        while True:
            try:
                # Display the current question
                print(f"\nQuestion #{question_count}")
                print(f"Current Question: {next_question}")

                # Get the response with validation
                answer = get_valid_input("Your Response: ", "answer")

                if answer.lower() == 'exit':
                    break

                # Process the response
                result = manager.process_response(next_question, answer)

                # Print analysis with improved formatting
                print("\n=== Analysis ===")
                print(f"Domain: {result['current_response']['domain']}")
                print(f"Confidence: {result['domain_confidence']:.2f}")
                print(f"Risk Level: {result['current_response']['risk_level']}")
                #print(f"Risk Score: {result['current_response']['risk_score']:.2f}")

                # Print recommendations if any
                if result['current_response']['recommendations']:
                    print("\n=== Recommendations ===")
                    for i, rec in enumerate(result['current_response']['recommendations'], 1):
                        print(f"{i}. {rec}")

                # Print reasoning
                print("\n=== Reasoning ===")
                print(result['current_response']['reasoning'])

                # Get the next question for the next iteration
                if result['next_question']:
                    next_question = result['next_question']
                    print("\n=== Next Question ===")
                    print(next_question)
                else:
                    # Fallback question if no next question is generated
                    next_question = "Can you describe any additional security measures or concerns?"

                print("\n" + "="*50)

                question_count += 1

            except Exception as e:
                print(f"\nError processing response: {str(e)}")
                print("Please try again.")
                continue

        # Generate and save final report
        if question_count > 1:  # Only generate report if there were questions
            try:
                final_report = manager.generate_report()

                # Save report with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_filename = f'security_assessment_report_{timestamp}.json'

                with open(report_filename, 'w') as f:
                    json.dump(final_report, f, indent=4)

                print(f"\nFinal Report generated and saved to '{report_filename}'")

                # Print summary
                print("\n=== Summary of Findings ===")
                print(f"Overall Risk Score: {final_report['assessment_summary']['overall_risk_score']:.2f}")
                print("\nDomain Risk Levels:")
                print("-" * 30)
                for domain, scores in final_report['assessment_summary']['domain_scores'].items():
                    risk_level = scores['risk_level']
                    risk_score = scores['score']
                    print(f"{domain:20}: {risk_level:8} ({risk_score:.2f})")

                # Print key recommendations
                print("\nKey Recommendations by Domain:")
                print("-" * 30)
                for domain, recs in final_report['recommendations'].items():
                    if recs:  # Only print domains with recommendations
                        print(f"\n{domain}:")
                        for i, rec in enumerate(recs[:3], 1):  # Show top 3 recommendations per domain
                            print(f"{i}. {rec}")

            except Exception as e:
                print(f"\nError generating report: {str(e)}")
                print("Please check the application logs for more details.")

        else:
            print("\nNo assessment data collected. Report generation skipped.")

    except KeyboardInterrupt:
        print("\n\nAssessment interrupted by user.")
        sys.exit(0)

    except Exception as e:
        print(f"\nCritical error: {str(e)}")
        print("The assessment has been terminated due to an unexpected error.")
        sys.exit(1)

if __name__ == "__main__":
    run_security_assessment()



# Load the JSON data from final.json
with open('/content/security_assessment_report_20241020_050008.json', 'r') as json_file:
    data = json.load(json_file)

wb = Workbook()
ws = wb.active
ws.title = "Risk Assessment Report"

# Define headers
headers = ['Question', 'Answer', 'Risk Level', 'Risk Score', 'Reasoning', 'Recommendations']
ws.append(headers)

# Define color fills for risk levels
fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # High risk
fill_orange = PatternFill(start_color="FFEBB5", end_color="FFEBB5", fill_type="solid")  # Medium risk
fill_green = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")  # Low risk

# Iterate through the conversation history in final.json
for entry in data['conversation_history']:
    row = [
        entry['question'],
        entry['answer'],
        entry['risk_level'],
        f"{entry['risk_score']:.2f}",
        entry['reasoning'],
        ', '.join(entry['recommendations'])
    ]
    ws.append(row)

    # Apply the fill color based on risk level
    last_row = ws.max_row  # Get the last row index
    if entry['risk_level'] == 'High':
        for cell in ws[last_row]:
            cell.fill = fill_red
    elif entry['risk_level'] == 'Medium':
        for cell in ws[last_row]:
            cell.fill = fill_orange
    elif entry['risk_level'] == 'Low':
        for cell in ws[last_row]:
            cell.fill = fill_green

# Save the workbook to a file
excel_file_name = "risk_assessment_report.xlsx"
wb.save(excel_file_name)
print(f"Excel report generated: {excel_file_name}")

